"""
FastAPI backend – REST API for the web dashboard.

Endpoints
---------
GET  /api/status              – portfolio snapshot (live prices)
GET  /api/history             – last N rebalance operations
GET  /api/snapshots           – portfolio value over time (for line chart)
GET  /api/config              – current config (assets, mode, settings)
POST /api/config              – update config (assets, allocations, mode…)
POST /api/rebalance           – trigger manual rebalance (returns job_id)
POST /api/rebalance/cancel    – cancel a pending rebalance within 10 s
GET  /api/export/csv          – download CSV report
GET  /api/export/excel        – download Excel report
GET  /api/bot/status          – rebalancer loop status
POST /api/bot/start           – start rebalancer loop
POST /api/bot/stop            – stop rebalancer loop
GET  /api/notifications/config – Discord / Telegram notification settings
POST /api/notifications/config – update notification settings
"""

import io
import csv
import os
import sys
import threading
import time
import logging
import uuid
import secrets
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Optional

_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _root not in sys.path:
    sys.path.insert(0, _root)

log = logging.getLogger("api")

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from database import (
    get_rebalance_history, get_snapshots, init_db, record_snapshot,
    save_portfolio, list_portfolios, get_portfolio,
    set_active_portfolio, delete_portfolio, update_portfolio_config,
    set_bot_running, get_running_portfolios,
    list_grid_bots, get_grid_bot, delete_grid_bot, get_grid_orders,
    get_should_run_grid_bots,
)
from grid_bot import (
    start_grid_bot, stop_grid_bot, resume_grid_bot,
    get_grid_bot_status, is_running as grid_is_running,
    calculate_grid_range, calculate_grid_count,
)
from mexc_client import MEXCClient
from smart_portfolio import (
    execute_rebalance,
    execute_rebalance_equal,
    get_pnl,
    get_portfolio_value,
    is_paper_trading,
    load_config,
    save_config,
    validate_allocations,
)

init_db()


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Per-portfolio rebalancer loop manager
# Each portfolio gets its own thread + stop event so multiple portfolios
# can run simultaneously and be stopped independently.
# ---------------------------------------------------------------------------

# portfolio_id -> {"thread": Thread, "stop": Event, "error": str|None, "started_at": str|None}
_portfolio_loops: dict[int, dict] = {}
_loops_lock = threading.Lock()


def _make_loop(portfolio_id: int, stop_event: threading.Event) -> None:
    """Rebalancer loop for a single portfolio."""
    from smart_portfolio import (
        execute_rebalance, needs_rebalance_proportional,
        next_run_time, get_portfolio_value, check_sl_tp,
        TIMED_FREQUENCY_MINUTES,
    )
    from database import get_portfolio as db_get_portfolio

    with _loops_lock:
        if portfolio_id in _portfolio_loops:
            _portfolio_loops[portfolio_id]["error"] = None
            _portfolio_loops[portfolio_id]["started_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    try:
        cfg = db_get_portfolio(portfolio_id)
        if cfg is None:
            raise ValueError(f"Portfolio {portfolio_id} not found")
        client = _client()
        mode = cfg["rebalance"]["mode"]
        log.info("Portfolio %d loop started | mode: %s", portfolio_id, mode)

        # ── Unified loop: re-reads cfg (including mode) from DB every cycle ──
        # This means mode changes, interval changes, and SL/TP config changes
        # all take effect on the next cycle without restarting the bot.
        timed_next_run = None  # initialised lazily on first timed cycle

        while not stop_event.is_set():
            try:
                cfg = db_get_portfolio(portfolio_id)
                if cfg is None:
                    log.error("Portfolio %d disappeared from DB — stopping loop", portfolio_id)
                    break

                current_mode = cfg["rebalance"]["mode"]

                # SL/TP guard runs on every cycle regardless of mode
                check_sl_tp(client, cfg)

                if current_mode == "proportional":
                    interval = cfg["rebalance"]["proportional"]["check_interval_minutes"] * 60
                    if needs_rebalance_proportional(client, cfg):
                        execute_rebalance(client, cfg)
                    stop_event.wait(interval)

                elif current_mode == "timed":
                    timed_cfg = cfg["rebalance"]["timed"]
                    frequency = timed_cfg["frequency"]
                    target_hour = timed_cfg.get("hour", 0)
                    # Initialise next_run on first timed cycle or after mode switch
                    if timed_next_run is None:
                        timed_next_run = next_run_time(frequency, target_hour=target_hour)
                        log.info("Portfolio %d: next timed rebalance at %s UTC", portfolio_id, timed_next_run.isoformat())
                    if datetime.utcnow() >= timed_next_run:
                        execute_rebalance(client, cfg)
                        timed_next_run = next_run_time(frequency, target_hour=target_hour)
                        log.info("Portfolio %d: next timed rebalance at %s UTC", portfolio_id, timed_next_run.isoformat())
                    short_freq = frequency in TIMED_FREQUENCY_MINUTES and frequency not in ("daily", "weekly", "monthly")
                    stop_event.wait(30 if short_freq else 60)

                else:  # unbalanced — SL/TP only
                    timed_next_run = None  # reset if mode switches back to timed later
                    stop_event.wait(60)

            except Exception as e:
                log.error("Portfolio %d loop error: %s", portfolio_id, e)
                stop_event.wait(30)  # back-off on error

    except Exception as e:
        with _loops_lock:
            if portfolio_id in _portfolio_loops:
                _portfolio_loops[portfolio_id]["error"] = str(e)
        log.error("Portfolio %d loop crashed: %s", portfolio_id, e)

    log.info("Portfolio %d loop stopped", portfolio_id)


def _is_portfolio_running(portfolio_id: int) -> bool:
    with _loops_lock:
        entry = _portfolio_loops.get(portfolio_id)
    return entry is not None and entry["thread"].is_alive()


def _start_portfolio_loop(portfolio_id: int) -> None:
    with _loops_lock:
        # Guard against double-start: if a live thread already exists, do nothing.
        existing = _portfolio_loops.get(portfolio_id)
        if existing is not None and existing["thread"].is_alive():
            log.warning("Portfolio %d loop already running — ignoring duplicate start", portfolio_id)
            return
        stop_ev = threading.Event()
        t = threading.Thread(target=_make_loop, args=(portfolio_id, stop_ev), daemon=True)
        _portfolio_loops[portfolio_id] = {
            "thread": t, "stop": stop_ev,
            "error": None, "started_at": None,
        }
    t.start()
    set_bot_running(portfolio_id, True)


def _stop_portfolio_loop(portfolio_id: int) -> None:
    with _loops_lock:
        entry = _portfolio_loops.get(portfolio_id)
    if entry:
        entry["stop"].set()
        # Wait up to 5 s for the thread to finish so we don't return before
        # any in-progress rebalance order is complete.
        entry["thread"].join(timeout=5)
        # Clean up dead entry so the dict doesn't grow unboundedly.
        with _loops_lock:
            if portfolio_id in _portfolio_loops and not _portfolio_loops[portfolio_id]["thread"].is_alive():
                del _portfolio_loops[portfolio_id]
    set_bot_running(portfolio_id, False)


# Legacy single-bot helpers kept for /api/bot/* endpoints (Dashboard tab)
def _is_running() -> bool:
    return any(_is_portfolio_running(pid) for pid in list(_portfolio_loops.keys()))


# ---------------------------------------------------------------------------
# Pending rebalance cancel window (10 seconds)
# ---------------------------------------------------------------------------
# Maps job_id -> {"cancel": threading.Event, "done": threading.Event, "result": list | None}
_pending_rebalances: dict[str, dict] = {}
_pending_lock = threading.Lock()


def _run_rebalance_with_cancel(job_id: str, client: MEXCClient, cfg: dict) -> None:
    """Execute rebalance after a 10-second cancel window."""
    entry = _pending_rebalances.get(job_id)
    if not entry:
        return
    cancelled = entry["cancel"].wait(timeout=10)
    if cancelled:
        entry["result"] = None
        entry["done"].set()
        log.info("Rebalance %s cancelled by user", job_id)
        return
    try:
        # Use active portfolio id for history isolation
        try:
            from database import list_portfolios as _lp
            _active = next((p for p in _lp() if p.get("active")), None)
            _pid = _active["id"] if _active else None
        except Exception:
            _pid = None
        result = execute_rebalance(client, cfg)
        entry["result"] = result
    except Exception as e:
        entry["result"] = [{"error": str(e)}]
        log.error("Rebalance %s failed: %s", job_id, e)
    finally:
        entry["done"].set()
        # Clean up after 60 s
        def _cleanup():
            time.sleep(60)
            with _pending_lock:
                _pending_rebalances.pop(job_id, None)
        threading.Thread(target=_cleanup, daemon=True).start()


# ---------------------------------------------------------------------------
# Discord notifications
# ---------------------------------------------------------------------------

def _send_discord(webhook_url: str, message: str) -> None:
    """Fire-and-forget Discord webhook notification."""
    try:
        import requests as _req
        _req.post(webhook_url, json={"content": message}, timeout=5)
    except Exception as e:
        log.warning("Discord notification failed: %s", e)


# ---------------------------------------------------------------------------
# FastAPI app
# ---------------------------------------------------------------------------

@asynccontextmanager
async def lifespan(app):
    # ── Startup: build frontend in background so uvicorn accepts requests
    #    immediately even if npm install+build takes 30+ seconds ──
    threading.Thread(target=_ensure_frontend_built, daemon=True).start()

    # ── Startup: resume portfolio loops that were running before restart ──
    try:
        running_ids = get_running_portfolios()
        if running_ids:
            log.info("Auto-resuming %d portfolio loop(s) after restart: %s", len(running_ids), running_ids)
            for pid in running_ids:
                try:
                    cfg = get_portfolio(pid)
                    if cfg is None:
                        log.warning("Auto-resume: portfolio %d not found in DB — skipping", pid)
                        set_bot_running(pid, False)
                        continue
                    _start_portfolio_loop(pid)
                    log.info("Auto-resumed portfolio %d (%s)", pid, cfg.get("bot", {}).get("name", "?"))
                except Exception as e:
                    log.error("Auto-resume failed for portfolio %d: %s", pid, e)
        else:
            log.info("No portfolios flagged for auto-resume")
    except Exception as e:
        log.error("Lifespan startup error (portfolios): %s", e)

    # ── Startup: resume grid bots that were running before restart ──
    try:
        grid_ids = get_should_run_grid_bots()
        if grid_ids:
            log.info("Auto-resuming %d grid bot(s) after restart: %s", len(grid_ids), grid_ids)
            for gid in grid_ids:
                try:
                    resume_grid_bot(gid)
                    log.info("Auto-resumed grid bot %d", gid)
                except Exception as e:
                    log.error("Auto-resume failed for grid bot %d: %s", gid, e)
        else:
            log.info("No grid bots flagged for auto-resume")
    except Exception as e:
        log.error("Lifespan startup error (grid bots): %s", e)

    yield
    # ── Shutdown: signal all loops to stop (Railway sends SIGTERM) ──
    log.info("Shutting down — stopping all portfolio loops")
    for pid in list(_portfolio_loops.keys()):
        try:
            _stop_portfolio_loop(pid)
        except Exception:
            pass


app = FastAPI(title="MEXC Portfolio Rebalancer API", version="3.0.1", lifespan=lifespan)


def _allowed_origins() -> list[str]:
    """Resolve allowed CORS origins from env.

    CORS_ALLOW_ORIGINS supports comma-separated origins.
    Defaults to localhost for safer local development.
    """
    raw = os.environ.get("CORS_ALLOW_ORIGINS", "").strip()
    if not raw:
        return [
            "http://localhost:3000",
            "http://127.0.0.1:3000",
            "http://localhost:8000",
            "http://127.0.0.1:8000",
        ]
    return [o.strip() for o in raw.split(",") if o.strip()]


app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins(),
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-API-Key"],
)

from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request as StarletteRequest
from starlette.responses import Response as StarletteResponse, JSONResponse

class NoCacheHtmlMiddleware(BaseHTTPMiddleware):
    """
    Force no-cache on every HTML response so Railway CDN and browsers
    never serve a stale build.  Also sets Surrogate-Control so Railway's
    edge layer bypasses its own cache for HTML.
    """
    async def dispatch(self, request: StarletteRequest, call_next):
        response: StarletteResponse = await call_next(request)
        ct = response.headers.get("content-type", "")
        if "text/html" in ct:
            response.headers["Cache-Control"]   = "no-cache, no-store, must-revalidate"
            response.headers["Pragma"]          = "no-cache"
            response.headers["Expires"]         = "0"
            response.headers["Surrogate-Control"] = "no-store"
            response.headers["Vary"]            = "*"
        return response

app.add_middleware(NoCacheHtmlMiddleware)


class ApiKeyAuthMiddleware(BaseHTTPMiddleware):
    """Protect sensitive /api/* routes using a shared API key.

    Configure with API_AUTH_KEY environment variable.
    Client can pass key in either:
      - Authorization: Bearer <key>
      - X-API-Key: <key>

    If API_AUTH_KEY is empty, auth is disabled (dev fallback).
    """

    _PUBLIC_PREFIXES = (
        "/health",
        "/docs",
        "/openapi.json",
        "/redoc",
        "/api/status",
        "/api/history",
        "/api/snapshots",
        "/api/config",
        "/api/rebalance/status/",
        "/api/mexc/status",
        "/api/db/status",
    )

    def _is_public(self, path: str) -> bool:
        if not path.startswith("/api") and path not in ("/health", "/docs", "/openapi.json", "/redoc"):
            return True
        return any(path == p or path.startswith(p) for p in self._PUBLIC_PREFIXES)

    async def dispatch(self, request: StarletteRequest, call_next):
        key = os.environ.get("API_AUTH_KEY", "").strip()
        if not key:
            return await call_next(request)

        path = request.url.path
        if self._is_public(path):
            return await call_next(request)

        auth = request.headers.get("authorization", "")
        x_api = request.headers.get("x-api-key", "")

        provided = ""
        if auth.lower().startswith("bearer "):
            provided = auth[7:].strip()
        elif x_api:
            provided = x_api.strip()

        if not provided or not secrets.compare_digest(provided, key):
            return JSONResponse(status_code=401, content={"detail": "Unauthorized"})

        return await call_next(request)


app.add_middleware(ApiKeyAuthMiddleware)

_static_dir = os.path.join(_root, "static")

# ── Auto-build frontend if static/index.html is missing ──────────────────────
def _ensure_frontend_built() -> None:
    """Build the Next.js frontend if the static output is absent.

    This runs once at import time so Railway deployments that skip the
    nixpacks build phase (e.g. restart-only deploys) still serve the UI.
    """
    index = os.path.join(_static_dir, "index.html")
    if os.path.exists(index):
        return
    web_dir = os.path.join(_root, "web")
    if not os.path.isdir(web_dir):
        log.warning("web/ directory not found — skipping frontend build")
        return
    import shutil
    import subprocess
    npm = shutil.which("npm")
    if npm is None:
        log.warning(
            "static/index.html missing and npm is not available — "
            "the frontend was not built. Deploy via nixpacks to include the UI."
        )
        return
    log.info("static/index.html missing — building Next.js frontend…")
    try:
        subprocess.run([npm, "install"], cwd=web_dir, check=True,
                       capture_output=True)
        subprocess.run([npm, "run", "build"], cwd=web_dir, check=True,
                       capture_output=True)
        out_dir = os.path.join(web_dir, "out")
        if os.path.isdir(out_dir):
            os.makedirs(_static_dir, exist_ok=True)
            shutil.copytree(out_dir, _static_dir, dirs_exist_ok=True)
            log.info("Frontend built and copied to static/")
        else:
            log.error("npm run build succeeded but web/out/ not found")
    except subprocess.CalledProcessError as e:
        log.error("Frontend build failed: %s", e)

# _ensure_frontend_built() is called inside lifespan() so it does not block
# the import phase (which would delay uvicorn startup in CI / cold starts).

# Next.js hashed assets (_next/static/**) are immutable — cache 1 year.
# HTML files must never be cached so users always get the latest build.
_IMMUTABLE = "public, max-age=31536000, immutable"
_NO_CACHE   = "no-cache, no-store, must-revalidate"

# Unique token per process start — injected into HTML responses so the
# browser ETag changes on every Railway redeploy even if the file bytes
# are identical.
import hashlib as _hashlib
import time as _time
_DEPLOY_ID = _hashlib.md5(str(_time.time()).encode()).hexdigest()[:12]

# Mount Next.js static assets so /_next/static/* is served correctly
_next_dir = os.path.join(_static_dir, "_next")
if os.path.isdir(_next_dir):
    app.mount("/_next", StaticFiles(directory=_next_dir), name="nextjs_static")


def _html_response(path: str):
    """Read an HTML file and inject deploy_id so every redeploy busts the cache."""
    from fastapi.responses import HTMLResponse
    with open(path, "r", encoding="utf-8") as f:
        html = f.read()
    # Inject deploy stamp into <head> so ETag/content changes every deploy
    stamp = f'<meta name="x-deploy-id" content="{_DEPLOY_ID}">'
    html = html.replace("</head>", f"{stamp}</head>", 1)
    headers = {
        "Cache-Control":    _NO_CACHE,
        "Pragma":           "no-cache",
        "Expires":          "0",
        "Surrogate-Control":"no-store",
        "Vary":             "*",
        "X-Deploy-Id":      _DEPLOY_ID,
    }
    return HTMLResponse(content=html, headers=headers)


@app.get("/", include_in_schema=False)
def serve_dashboard():
    index = os.path.join(_static_dir, "index.html")
    if os.path.exists(index):
        return _html_response(index)
    return {"message": "MEXC Rebalancer API", "docs": "/docs"}


def _client() -> MEXCClient:
    return MEXCClient()


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class AssetAlloc(BaseModel):
    symbol: str
    allocation_pct: float
    entry_price_usdt: Optional[float] = None


class ConfigUpdate(BaseModel):
    bot_name: Optional[str] = None
    assets: Optional[list[AssetAlloc]] = None
    total_usdt: Optional[float] = None
    allocation_mode: Optional[str] = None          # ai_balance | equal | market_cap
    rebalance_mode: Optional[str] = None
    threshold_pct: Optional[int] = None
    frequency: Optional[str] = None
    timed_hour: Optional[int] = None
    sell_at_termination: Optional[bool] = None
    enable_asset_transfer: Optional[bool] = None
    paper_trading: Optional[bool] = None
    stop_loss_pct: Optional[float] = None          # 1–100, None = disabled
    take_profit_pct: Optional[float] = None        # 1–500, None = disabled


class NotificationConfig(BaseModel):
    discord_webhook_url: Optional[str] = None
    discord_enabled: Optional[bool] = None



# ---------------------------------------------------------------------------
# Routes – Status & portfolio
# ---------------------------------------------------------------------------

@app.get("/api/status")
def get_status():
    cfg = load_config()
    no_key = not os.environ.get("MEXC_API_KEY")
    if no_key:
        assets_out = [
            {
                "symbol": a["symbol"],
                "balance": 0,
                "price": 0,
                "price_usdt": 0,
                "value_usdt": 0,
                "actual_pct": a["allocation_pct"],
                "current_pct": a["allocation_pct"],
                "target_pct": a["allocation_pct"],
                "deviation": 0,
                "diff_pct": 0,
            }
            for a in cfg["portfolio"]["assets"]
        ]
        return {
            "bot_name": cfg["bot"]["name"],
            "total_usdt": cfg["portfolio"].get("total_usdt", 0),
            "mode": cfg["rebalance"]["mode"],
            "paper_trading": cfg.get("paper_trading", False),
            "last_rebalance": cfg.get("last_rebalance"),
            "assets": assets_out,
            "pnl": {"initial_usdt": 0, "current_usdt": 0, "pnl_usdt": 0, "pnl_pct": 0},
            "profit_usdt": 0,
            "profit_pct": 0,
            "warning": "MEXC API key not set — showing config values only",
            "rebalance_config": {
                "threshold_pct": cfg["rebalance"]["proportional"]["threshold_pct"],
                "frequency": cfg["rebalance"]["timed"]["frequency"],
            },
        }
    try:
        client = _client()
        # Status uses live values so the user sees real balances.
        # budget_usdt is only used during rebalance execution.
        portfolio = get_portfolio_value(client, cfg["portfolio"]["assets"], budget_usdt=None)
        targets = {a["symbol"]: a["allocation_pct"] for a in cfg["portfolio"]["assets"]}
        pnl = get_pnl(cfg, current_usdt=portfolio["total_usdt"])
        assets_out = []
        for r in portfolio["assets"]:
            diff = round(r["actual_pct"] - targets[r["symbol"]], 2)
            assets_out.append({
                "symbol": r["symbol"],
                "balance": r["balance"],
                "price": r["price"],
                "price_usdt": r["price"],
                "value_usdt": r["value_usdt"],
                "actual_pct": round(r["actual_pct"], 2),
                "current_pct": round(r["actual_pct"], 2),
                "target_pct": targets[r["symbol"]],
                "deviation": diff,
                "diff_pct": diff,
                "error": r.get("error"),
            })
        invalid = portfolio.get("invalid_symbols", [])
        response = {
            "bot_name": cfg["bot"]["name"],
            "total_usdt": portfolio["total_usdt"],
            "mode": cfg["rebalance"]["mode"],
            "paper_trading": cfg.get("paper_trading", False),
            "last_rebalance": cfg.get("last_rebalance"),
            "assets": assets_out,
            "pnl": pnl,
            "profit_usdt": pnl.get("pnl_usdt", 0),
            "profit_pct": pnl.get("pnl_pct", 0),
            "rebalance_config": {
                "threshold_pct": cfg["rebalance"]["proportional"]["threshold_pct"],
                "frequency": cfg["rebalance"]["timed"]["frequency"],
            },
        }
        if invalid:
            response["warning"] = f"رموز غير موجودة على MEXC: {', '.join(invalid)} — تحقق من الإعدادات"
        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/history")
def get_history(limit: int = 10):
    return get_rebalance_history(limit)


@app.get("/api/snapshots")
def get_portfolio_snapshots(limit: int = 90):
    return get_snapshots(limit)


# ---------------------------------------------------------------------------
# Routes – Config
# ---------------------------------------------------------------------------

@app.get("/api/config")
def get_config():
    return load_config()


@app.post("/api/config")
def update_config(body: ConfigUpdate):
    cfg = load_config()
    if body.bot_name is not None:
        cfg["bot"]["name"] = body.bot_name
    if body.assets is not None:
        symbols = [a.symbol.strip().upper() for a in body.assets]
        if len(symbols) != len(set(symbols)):
            raise HTTPException(status_code=400, detail="لا يمكن تكرار رموز العملات")
        # Validate symbols exist on MEXC (only when API key is set)
        if os.environ.get("MEXC_API_KEY"):
            try:
                client = _client()
                invalid = []
                for sym in symbols:
                    if sym == "USDT":
                        continue
                    try:
                        price = client.get_price(f"{sym}USDT")
                        if price <= 0:
                            invalid.append(sym)
                    except Exception:
                        invalid.append(sym)
                if invalid:
                    raise HTTPException(
                        status_code=400,
                        detail=f"الرموز التالية غير موجودة على MEXC: {', '.join(invalid)}"
                    )
            except HTTPException:
                raise
            except Exception:
                pass  # If validation itself fails, allow save
        cfg["portfolio"]["assets"] = [
            {
                "symbol": s,
                "allocation_pct": a.allocation_pct,
                "entry_price_usdt": a.entry_price_usdt,
            }
            for s, a in zip(symbols, body.assets)
        ]
    if body.allocation_mode is not None:
        cfg["portfolio"]["allocation_mode"] = body.allocation_mode
    if body.total_usdt is not None:
        cfg["portfolio"]["total_usdt"] = body.total_usdt
        # Always update initial baseline when user explicitly sets total_usdt
        cfg["portfolio"]["initial_value_usdt"] = body.total_usdt
    if body.rebalance_mode is not None:
        cfg["rebalance"]["mode"] = body.rebalance_mode
    if body.threshold_pct is not None:
        cfg["rebalance"]["proportional"]["threshold_pct"] = body.threshold_pct
    if body.frequency is not None:
        cfg["rebalance"]["timed"]["frequency"] = body.frequency
    if body.timed_hour is not None:
        cfg["rebalance"]["timed"]["hour"] = max(0, min(23, body.timed_hour))
    if body.sell_at_termination is not None:
        cfg["termination"]["sell_at_termination"] = body.sell_at_termination
    if body.enable_asset_transfer is not None:
        cfg["asset_transfer"]["enable_asset_transfer"] = body.enable_asset_transfer
    if body.paper_trading is not None:
        cfg["paper_trading"] = body.paper_trading
    # Risk settings
    if body.stop_loss_pct is not None or body.take_profit_pct is not None:
        if "risk" not in cfg:
            cfg["risk"] = {}
        if body.stop_loss_pct is not None:
            if body.stop_loss_pct == 0:
                cfg["risk"]["stop_loss_pct"] = None
            elif 1 <= body.stop_loss_pct <= 100:
                cfg["risk"]["stop_loss_pct"] = body.stop_loss_pct
            else:
                raise HTTPException(status_code=400, detail="stop_loss_pct must be 1–100 or 0 to disable")
        if body.take_profit_pct is not None:
            if body.take_profit_pct == 0:
                cfg["risk"]["take_profit_pct"] = None
            elif 1 <= body.take_profit_pct <= 500:
                cfg["risk"]["take_profit_pct"] = body.take_profit_pct
            else:
                raise HTTPException(status_code=400, detail="take_profit_pct must be 1–500 or 0 to disable")
    try:
        validate_allocations(cfg["portfolio"]["assets"])
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    save_config(cfg)
    # Mirror changes to the active portfolio's DB record so total_usdt stays isolated per portfolio
    try:
        portfolios_list = list_portfolios()
        active = next((p for p in portfolios_list if p.get("active")), None)
        if active:
            active_cfg = get_portfolio(active["id"])
            if active_cfg:
                if body.assets is not None:
                    active_cfg["portfolio"]["assets"] = cfg["portfolio"]["assets"]
                if body.total_usdt is not None:
                    active_cfg["portfolio"]["total_usdt"] = cfg["portfolio"]["total_usdt"]
                    active_cfg["portfolio"]["initial_value_usdt"] = cfg["portfolio"]["initial_value_usdt"]
                if body.rebalance_mode is not None:
                    active_cfg["rebalance"]["mode"] = cfg["rebalance"]["mode"]
                if body.threshold_pct is not None:
                    active_cfg["rebalance"]["proportional"]["threshold_pct"] = cfg["rebalance"]["proportional"]["threshold_pct"]
                if body.frequency is not None:
                    active_cfg["rebalance"]["timed"]["frequency"] = cfg["rebalance"]["timed"]["frequency"]
                if body.timed_hour is not None:
                    active_cfg["rebalance"]["timed"]["hour"] = cfg["rebalance"]["timed"]["hour"]
                if body.paper_trading is not None:
                    active_cfg["paper_trading"] = cfg["paper_trading"]
                if body.allocation_mode is not None:
                    active_cfg["portfolio"]["allocation_mode"] = cfg["portfolio"]["allocation_mode"]
                if "risk" in cfg:
                    active_cfg["risk"] = cfg["risk"]
                update_portfolio_config(active["id"], active_cfg)
    except Exception as e:
        log.warning("update_config: could not mirror to active portfolio DB record: %s", e)
    return {"ok": True}


@app.get("/api/account/total")
def get_account_total():
    """Return total account value across ALL assets on MEXC (not just portfolio)."""
    try:
        client = _client()
        balances = client.get_spot_assets()
        total = 0.0
        assets_out = []
        for b in balances:
            sym = b.get("asset", "").upper()
            free = float(b.get("free", 0))
            locked = float(b.get("locked", 0))
            qty = free + locked
            if qty <= 0:
                continue
            if sym == "USDT":
                price = 1.0
            else:
                try:
                    price = client.get_price(f"{sym}USDT")
                except Exception:
                    continue  # skip assets with no USDT pair
            value = qty * price
            if value < 0.01:
                continue
            total += value
            assets_out.append({
                "symbol": sym,
                "balance": qty,
                "price_usdt": price,
                "value_usdt": round(value, 4),
            })
        assets_out.sort(key=lambda x: x["value_usdt"], reverse=True)
        # Compute free vs locked totals across all assets
        free_total   = 0.0
        locked_total = 0.0
        for b in balances:
            sym    = b.get("asset", "").upper()
            free   = float(b.get("free",   0))
            locked = float(b.get("locked", 0))
            if free + locked <= 0:
                continue
            if sym == "USDT":
                price = 1.0
            else:
                try:
                    price = client.get_price(f"{sym}USDT")
                except Exception:
                    continue
            free_total   += free   * price
            locked_total += locked * price
        return {
            "ok": True,
            "total_usdt":  round(total, 2),
            "free_usdt":   round(free_total,   2),
            "locked_usdt": round(locked_total, 2),
            "assets": assets_out,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/config/reset-initial-value")
def reset_initial_value():
    """Reset initial_value_usdt for the active portfolio to the current live MEXC value."""
    # Operate on the active portfolio in DB so each portfolio's total_usdt stays isolated
    portfolios = list_portfolios()
    active = next((p for p in portfolios if p.get("active")), None)
    if active:
        cfg = get_portfolio(active["id"])
        portfolio_id = active["id"]
    else:
        cfg = load_config()
        portfolio_id = None
    if cfg is None:
        raise HTTPException(status_code=404, detail="لا توجد محفظة نشطة")
    try:
        client = _client()
        portfolio = get_portfolio_value(client, cfg["portfolio"]["assets"], budget_usdt=None)
        live_total = portfolio["total_usdt"]
        if live_total <= 0:
            raise HTTPException(status_code=400, detail="القيمة الحالية صفر — تحقق من الرصيد")
        cfg["portfolio"]["total_usdt"] = round(live_total, 2)
        cfg["portfolio"]["initial_value_usdt"] = round(live_total, 2)
        if portfolio_id is not None:
            update_portfolio_config(portfolio_id, cfg)
        else:
            save_config(cfg)
        return {"ok": True, "initial_value_usdt": round(live_total, 2)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Routes – Rebalance (with 10-second cancel window)
# ---------------------------------------------------------------------------

@app.post("/api/rebalance")
def trigger_rebalance():
    """
    Starts a rebalance with a 10-second cancel window.
    Returns a job_id that can be passed to /api/rebalance/cancel.
    """
    cfg = load_config()
    try:
        client = _client()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    job_id = str(uuid.uuid4())
    entry = {
        "cancel": threading.Event(),
        "done": threading.Event(),
        "result": None,
    }
    with _pending_lock:
        _pending_rebalances[job_id] = entry

    t = threading.Thread(
        target=_run_rebalance_with_cancel,
        args=(job_id, client, cfg),
        daemon=True,
    )
    t.start()

    return {"ok": True, "job_id": job_id, "cancel_window_seconds": 10}


@app.post("/api/rebalance/cancel")
def cancel_rebalance(job_id: str):
    """Cancel a pending rebalance within the 10-second window."""
    with _pending_lock:
        entry = _pending_rebalances.get(job_id)
    if not entry:
        raise HTTPException(status_code=404, detail="لم يتم العثور على العملية أو انتهت مهلة الإلغاء")
    if entry["done"].is_set():
        raise HTTPException(status_code=409, detail="تم تنفيذ العملية بالفعل ولا يمكن إلغاؤها")
    entry["cancel"].set()
    return {"ok": True, "message": "تم إلغاء عملية Rebalance"}


@app.get("/api/rebalance/status/{job_id}")
def rebalance_job_status(job_id: str):
    """Poll the status of a rebalance job."""
    with _pending_lock:
        entry = _pending_rebalances.get(job_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Job not found")
    cancelled = entry["cancel"].is_set()
    done = entry["done"].is_set()
    return {
        "job_id": job_id,
        "cancelled": cancelled,
        "done": done,
        "result": entry.get("result") if done else None,
    }


# ---------------------------------------------------------------------------
# Routes – Notifications
# ---------------------------------------------------------------------------

@app.get("/api/notifications/config")
def get_notification_config():
    cfg = load_config()
    notif = cfg.get("notifications", {})
    return {
        "discord_enabled": notif.get("discord_enabled", False),
        "discord_webhook_url": notif.get("discord_webhook_url", ""),
        "telegram_enabled": False,
    }


@app.post("/api/notifications/config")
def update_notification_config(body: NotificationConfig):
    cfg = load_config()
    if "notifications" not in cfg:
        cfg["notifications"] = {}
    if body.discord_webhook_url is not None:
        cfg["notifications"]["discord_webhook_url"] = body.discord_webhook_url
    if body.discord_enabled is not None:
        cfg["notifications"]["discord_enabled"] = body.discord_enabled

    save_config(cfg)
    return {"ok": True}


@app.post("/api/notifications/test")
def test_discord_notification():
    cfg = load_config()
    notif = cfg.get("notifications", {})
    webhook = notif.get("discord_webhook_url", "")
    if not webhook:
        raise HTTPException(status_code=400, detail="Discord webhook URL غير مضبوط")
    _send_discord(webhook, "✅ اختبار إشعار Discord من MEXC Smart Portfolio Bot")
    return {"ok": True}


# ---------------------------------------------------------------------------
# Routes – Export
# ---------------------------------------------------------------------------

@app.get("/api/export/csv")
def export_csv():
    rows = get_rebalance_history(500)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["timestamp", "mode", "total_usdt", "paper",
                "symbol", "target_pct", "actual_pct",
                "deviation", "diff_usdt", "action"])
    for r in rows:
        for d in r["details"]:
            w.writerow([
                r["ts"], r["mode"], r["total_usdt"], bool(r["paper"]),
                d["symbol"], d["target_pct"], d["actual_pct"],
                d["deviation"], d["diff_usdt"], d["action"],
            ])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=rebalance_history.csv"},
    )


@app.get("/api/export/excel")
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl غير مثبت")

    rows = get_rebalance_history(500)
    wb = openpyxl.Workbook()

    # ── Sheet 1: Rebalance History ──────────────────────────────────────────
    ws = wb.active
    ws.title = "سجل العمليات"

    headers = ["الوقت", "الوضع", "الإجمالي (USDT)", "تجريبي",
               "العملة", "الهدف%", "الحالي%", "الفرق%", "الفرق (USDT)", "الإجراء"]
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="F0B90B")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    for r in rows:
        for d in r["details"]:
            action = d.get("action", "")
            ws.cell(row=row_num, column=1, value=r["ts"])
            ws.cell(row=row_num, column=2, value=r["mode"])
            ws.cell(row=row_num, column=3, value=round(r["total_usdt"], 2))
            ws.cell(row=row_num, column=4, value="نعم" if r["paper"] else "لا")
            ws.cell(row=row_num, column=5, value=d.get("symbol", ""))
            ws.cell(row=row_num, column=6, value=d.get("target_pct", 0))
            ws.cell(row=row_num, column=7, value=d.get("actual_pct", 0))
            ws.cell(row=row_num, column=8, value=d.get("deviation", 0))
            ws.cell(row=row_num, column=9, value=d.get("diff_usdt", 0))
            action_cell = ws.cell(row=row_num, column=10, value=action)
            if action == "BUY":
                action_cell.font = Font(color="10B981", bold=True)
            elif action == "SELL":
                action_cell.font = Font(color="EF4444", bold=True)
            row_num += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet 2: Portfolio Snapshots ────────────────────────────────────────
    ws2 = wb.create_sheet("أداء المحفظة")
    ws2.cell(row=1, column=1, value="الوقت").font = header_font
    ws2.cell(row=1, column=2, value="الإجمالي (USDT)").font = header_font
    for i, snap in enumerate(get_snapshots(500), 2):
        ws2.cell(row=i, column=1, value=snap["ts"])
        ws2.cell(row=i, column=2, value=round(snap["total_usdt"], 2))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"portfolio_report_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# Routes – Health & bot control
# ---------------------------------------------------------------------------

@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/api/bot/status")
def bot_status():
    """Overall status — aggregates all running portfolio loops."""
    with _loops_lock:
        loops_snapshot = {pid: dict(e) for pid, e in _portfolio_loops.items()}
    running_ids = [pid for pid, e in loops_snapshot.items() if e["thread"].is_alive()]
    return {
        "running": len(running_ids) > 0,
        "running_portfolios": running_ids,
        "started_at": next((loops_snapshot[pid]["started_at"] for pid in running_ids), None),
        "error": next((loops_snapshot[pid]["error"] for pid in running_ids if loops_snapshot[pid]["error"]), None),
        "mode": load_config()["rebalance"]["mode"],
    }


@app.post("/api/bot/start")
def bot_start():
    """Start the active portfolio's loop (legacy single-bot endpoint)."""
    cfg = load_config()
    try:
        validate_allocations(cfg["portfolio"]["assets"])
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    # Find the active portfolio id
    portfolios = list_portfolios()
    active = next((p for p in portfolios if p.get("active")), None)
    if active is None:
        raise HTTPException(status_code=400, detail="لا توجد محفظة نشطة")
    pid = active["id"]
    if _is_portfolio_running(pid):
        return {"ok": False, "message": "البوت شغال بالفعل"}
    _start_portfolio_loop(pid)
    return {"ok": True, "message": f"البوت بدأ | mode: {cfg['rebalance']['mode']}"}


@app.post("/api/bot/stop")
def bot_stop():
    """Stop all running portfolio loops."""
    running = [pid for pid in list(_portfolio_loops.keys()) if _is_portfolio_running(pid)]
    if not running:
        return {"ok": False, "message": "البوت مش شغال"}
    for pid in running:
        _stop_portfolio_loop(pid)
    return {"ok": True, "message": "تم إيقاف البوت"}


# ---------------------------------------------------------------------------
# Routes – Multi-portfolio management
# ---------------------------------------------------------------------------

class PortfolioCreate(BaseModel):
    config: dict


@app.get("/api/portfolios")
def api_list_portfolios():
    """List all saved portfolios with live running state."""
    portfolios = list_portfolios()
    for p in portfolios:
        p["running"] = _is_portfolio_running(p["id"])
    return portfolios


@app.post("/api/portfolios")
def api_save_portfolio(body: PortfolioCreate):
    """Save current config as a new named portfolio."""
    cfg = body.config
    try:
        validate_allocations(cfg["portfolio"]["assets"])
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=str(e))
    name = cfg.get("bot", {}).get("name", "Portfolio")
    try:
        pid = save_portfolio(name, cfg)
    except Exception as e:
        log.error("save_portfolio exception: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"فشل حفظ المحفظة: {type(e).__name__}: {e}")
    return {"ok": True, "id": pid}


@app.get("/api/db/status")
def db_status():
    """Diagnostic endpoint — returns DB backend and connection health."""
    import database as _db
    backend = "postgresql" if _db._USE_POSTGRES else "sqlite"
    try:
        count = len(_db.list_portfolios())
        return {"backend": backend, "ok": True, "portfolio_count": count}
    except Exception as e:
        return {"backend": backend, "ok": False, "error": str(e)}


@app.get("/api/mexc/status")
def mexc_status():
    """Diagnostic endpoint — checks MEXC API key, connectivity, and balances."""
    api_key = os.environ.get("MEXC_API_KEY", "")
    secret  = os.environ.get("MEXC_SECRET_KEY", "")
    if not api_key or not secret:
        return {"ok": False, "error": "MEXC_API_KEY أو MEXC_SECRET_KEY غير موجود في متغيرات البيئة"}
    try:
        client = MEXCClient(api_key, secret)
        account = client.get_account()
        all_balances = account.get("balances", [])
        non_zero = [
            b for b in all_balances
            if float(b.get("free", 0)) > 0 or float(b.get("locked", 0)) > 0
        ]
        # Check balances for configured portfolio assets
        cfg = load_config()
        portfolio_assets = cfg.get("portfolio", {}).get("assets", [])
        portfolio_balances = []
        for a in portfolio_assets:
            sym = a["symbol"]
            bal = next((b for b in all_balances if b["asset"] == sym), None)
            try:
                price = client.get_price(f"{sym}USDT") if sym != "USDT" else 1.0
            except Exception:
                price = 0.0
            free = float(bal["free"]) if bal else 0.0
            portfolio_balances.append({
                "symbol": sym,
                "free": free,
                "price_usdt": price,
                "value_usdt": round(free * price, 2),
                "target_pct": a["allocation_pct"],
            })
        return {
            "ok": True,
            "key_prefix": api_key[:6] + "...",
            "paper_trading": cfg.get("paper_trading", False),
            "total_non_zero_assets": len(non_zero),
            "all_non_zero": [{"asset": b["asset"], "free": b["free"]} for b in non_zero[:20]],
            "portfolio_assets": portfolio_balances,
            "total_portfolio_usdt": round(sum(b["value_usdt"] for b in portfolio_balances), 2),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/portfolios/{portfolio_id}")
def api_get_portfolio(portfolio_id: int):
    cfg = get_portfolio(portfolio_id)
    if cfg is None:
        raise HTTPException(status_code=404, detail="المحفظة غير موجودة")
    return cfg


@app.post("/api/portfolios/{portfolio_id}/activate")
def api_activate_portfolio(portfolio_id: int):
    """Mark portfolio as the active one in DB. Each portfolio keeps its own total_usdt."""
    cfg = get_portfolio(portfolio_id)
    if cfg is None:
        raise HTTPException(status_code=404, detail="المحفظة غير موجودة")
    try:
        validate_allocations(cfg["portfolio"]["assets"])
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    # Snapshot real portfolio value from MEXC and persist to this portfolio's DB record only
    try:
        client = _client()
        portfolio = get_portfolio_value(client, cfg["portfolio"]["assets"], budget_usdt=None)
        live_total = portfolio["total_usdt"]
        if live_total > 0:
            cfg["portfolio"]["total_usdt"] = round(live_total, 2)
            cfg["portfolio"]["initial_value_usdt"] = round(live_total, 2)
            update_portfolio_config(portfolio_id, cfg)
    except Exception as e:
        log.warning("activate: could not fetch live value: %s", e)
    set_active_portfolio(portfolio_id)
    return {"ok": True, "message": f"تم تفعيل المحفظة: {cfg['bot']['name']}"}


@app.post("/api/portfolios/{portfolio_id}/start")
def api_start_portfolio(portfolio_id: int):
    """Start the rebalancer loop for a specific portfolio. total_usdt is isolated per portfolio in DB."""
    cfg = get_portfolio(portfolio_id)
    if cfg is None:
        raise HTTPException(status_code=404, detail="المحفظة غير موجودة")
    try:
        validate_allocations(cfg["portfolio"]["assets"])
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if _is_portfolio_running(portfolio_id):
        return {"ok": False, "message": "المحفظة شغالة بالفعل"}
    # Sync real portfolio value before starting — persist to this portfolio's DB record only
    try:
        client = _client()
        portfolio = get_portfolio_value(client, cfg["portfolio"]["assets"], budget_usdt=None)
        live_total = portfolio["total_usdt"]
        if live_total > 0:
            cfg["portfolio"]["total_usdt"] = round(live_total, 2)
            if "initial_value_usdt" not in cfg["portfolio"] or cfg["portfolio"].get("initial_value_usdt", 0) <= 0:
                cfg["portfolio"]["initial_value_usdt"] = round(live_total, 2)
            update_portfolio_config(portfolio_id, cfg)
    except Exception as e:
        log.warning("start: could not fetch live value: %s", e)
    _start_portfolio_loop(portfolio_id)
    return {"ok": True, "message": f"تم تشغيل المحفظة: {cfg['bot']['name']}"}


@app.post("/api/portfolios/{portfolio_id}/stop")
def api_stop_portfolio(portfolio_id: int):
    """Stop the rebalancer loop for a specific portfolio."""
    if not _is_portfolio_running(portfolio_id):
        return {"ok": False, "message": "المحفظة مش شغالة"}
    _stop_portfolio_loop(portfolio_id)
    return {"ok": True, "message": "تم إيقاف المحفظة"}


@app.get("/api/portfolios/{portfolio_id}/status")
def api_portfolio_status(portfolio_id: int):
    """Return running state for a specific portfolio."""
    with _loops_lock:
        entry = _portfolio_loops.get(portfolio_id)
    running = entry is not None and entry["thread"].is_alive()
    return {
        "portfolio_id": portfolio_id,
        "running": running,
        "started_at": entry["started_at"] if entry else None,
        "error": entry["error"] if entry else None,
    }


@app.delete("/api/portfolios/{portfolio_id}")
def api_delete_portfolio(portfolio_id: int):
    delete_portfolio(portfolio_id)
    return {"ok": True}


@app.put("/api/portfolios/{portfolio_id}")
def api_update_portfolio(portfolio_id: int, body: PortfolioCreate):
    """Update a saved portfolio's config without activating it."""
    cfg = body.config
    try:
        validate_allocations(cfg["portfolio"]["assets"])
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=str(e))
    existing = get_portfolio(portfolio_id)
    if existing is None:
        raise HTTPException(status_code=404, detail="المحفظة غير موجودة")
    update_portfolio_config(portfolio_id, cfg)
    return {"ok": True}


class PortfolioRebalanceRequest(BaseModel):
    rebalance_type: str = "market_value"  # "market_value" | "equal"


@app.post("/api/portfolios/{portfolio_id}/rebalance")
def api_rebalance_portfolio(portfolio_id: int, body: PortfolioRebalanceRequest):
    """
    Trigger a rebalance for a specific portfolio.

    rebalance_type:
      - "market_value": restore configured allocation_pct targets (default)
      - "equal": redistribute equally across all assets regardless of targets
    """
    cfg = get_portfolio(portfolio_id)
    if cfg is None:
        raise HTTPException(status_code=404, detail="المحفظة غير موجودة")
    try:
        client = _client()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    job_id = str(uuid.uuid4())
    entry: dict = {"cancel": threading.Event(), "done": threading.Event(), "result": None}
    with _pending_lock:
        _pending_rebalances[job_id] = entry

    def _run() -> None:
        cancel_ev: threading.Event = entry["cancel"]
        cancel_ev.wait(10)
        if cancel_ev.is_set():
            entry["done"].set()
            return
        try:
            if body.rebalance_type == "equal":
                result = execute_rebalance_equal(client, cfg)
            else:
                result = execute_rebalance(client, cfg)
            entry["result"] = result
        except Exception as exc:
            log.error("Portfolio rebalance failed (id=%s): %s", portfolio_id, exc)
            entry["result"] = [{"error": str(exc)}]
        finally:
            entry["done"].set()

    threading.Thread(target=_run, daemon=True).start()
    return {"ok": True, "job_id": job_id, "cancel_window_seconds": 10}


@app.post("/api/portfolios/{portfolio_id}/stop-and-sell")
def api_stop_and_sell(portfolio_id: int):
    """
    Stop the rebalancer loop (if running) then sell all portfolio assets to USDT.
    Only works on the active portfolio.
    """
    cfg = get_portfolio(portfolio_id)
    if cfg is None:
        raise HTTPException(status_code=404, detail="المحفظة غير موجودة")

    # Stop this portfolio's loop first
    if _is_portfolio_running(portfolio_id):
        _stop_portfolio_loop(portfolio_id)

    try:
        client = _client()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    paper = is_paper_trading(cfg)
    assets = cfg["portfolio"]["assets"]
    results = []
    for a in assets:
        sym = a["symbol"]
        if sym == "USDT":
            continue
        try:
            balance = client.get_asset_balance(sym)
            if balance <= 0:
                results.append({"symbol": sym, "action": "SKIP", "reason": "zero balance"})
                continue
            if paper:
                log.info("stop-and-sell [PAPER]: would sell %.8f %s", balance, sym)
                results.append({"symbol": sym, "action": "PAPER_SELL", "qty": balance})
            else:
                resp = client.place_market_sell(f"{sym}USDT", balance)
                log.info("stop-and-sell: sold %.8f %s → %s", balance, sym, resp)
                results.append({"symbol": sym, "action": "SOLD", "qty": balance})
        except Exception as e:
            log.error("stop-and-sell: failed to sell %s: %s", sym, e)
            results.append({"symbol": sym, "action": "ERROR", "error": str(e)})

    return {"ok": True, "paper": paper, "results": results}


@app.get("/api/portfolios/{portfolio_id}/assets")
def api_portfolio_assets(portfolio_id: int):
    """
    Return live asset breakdown for a specific portfolio (same shape as /api/status assets).
    Used by the dashboard portfolio selector to show per-portfolio coin rows.
    """
    cfg = get_portfolio(portfolio_id)
    if cfg is None:
        raise HTTPException(status_code=404, detail="المحفظة غير موجودة")

    no_key = not os.environ.get("MEXC_API_KEY")
    port_assets = cfg["portfolio"]["assets"]
    targets = {a["symbol"]: a["allocation_pct"] for a in port_assets}

    if no_key:
        assets_out = [
            {
                "symbol": a["symbol"],
                "balance": 0, "price": 0, "price_usdt": 0,
                "value_usdt": 0,
                "actual_pct": a["allocation_pct"],
                "current_pct": a["allocation_pct"],
                "target_pct": a["allocation_pct"],
                "deviation": 0, "diff_pct": 0,
            }
            for a in port_assets
        ]
        return {
            "portfolio_id": portfolio_id,
            "portfolio_name": cfg["bot"]["name"],
            "total_usdt": cfg["portfolio"].get("total_usdt", 0),
            "mode": cfg["rebalance"]["mode"],
            "running": _is_portfolio_running(portfolio_id),
            "assets": assets_out,
        }

    try:
        client = _client()
        portfolio = get_portfolio_value(client, port_assets, budget_usdt=None)
        assets_out = []
        for r in portfolio["assets"]:
            diff = round(r["actual_pct"] - targets.get(r["symbol"], 0), 2)
            assets_out.append({
                "symbol": r["symbol"],
                "balance": r["balance"],
                "price": r["price"],
                "price_usdt": r["price"],
                "value_usdt": r["value_usdt"],
                "actual_pct": round(r["actual_pct"], 2),
                "current_pct": round(r["actual_pct"], 2),
                "target_pct": targets.get(r["symbol"], 0),
                "deviation": diff,
                "diff_pct": diff,
                "error": r.get("error"),
            })
        return {
            "portfolio_id": portfolio_id,
            "portfolio_name": cfg["bot"]["name"],
            "total_usdt": portfolio["total_usdt"],
            "mode": cfg["rebalance"]["mode"],
            "running": _is_portfolio_running(portfolio_id),
            "assets": assets_out,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Grid Bot endpoints
# ---------------------------------------------------------------------------

class GridBotCreate(BaseModel):
    symbol: str
    investment: float
    grid_count: int | None = None
    price_low: float | None = None
    price_high: float | None = None
    mode: str = "normal"               # 'normal' | 'infinity'
    use_base_balance: bool = False     # include existing base-asset value


@app.get("/api/grid-bots")
def api_list_grid_bots():
    bots = list_grid_bots()
    result = []
    for b in bots:
        result.append({
            "id":              b["id"],
            "symbol":          b["symbol"],
            "investment":      b["investment"],
            "grid_count":      b["grid_count"],
            "price_low":       b["price_low"],
            "price_high":      b["price_high"],
            "mode":            b.get("mode", "normal"),
            "status":          b["status"],
            "profit":          round(float(b.get("realised_profit") or 0) + float(b.get("unrealized_pnl") or 0), 4),
            "unrealized_pnl":  b.get("unrealized_pnl", 0),
            "realised_profit": b.get("realised_profit", 0),
            "avg_buy_price":   b.get("avg_buy_price", 0),
            "base_qty":        b.get("base_qty", 0),
            "running":         grid_is_running(b["id"]),
            "ts_created":      b["ts_created"],
        })
    return result


@app.post("/api/grid-bots")
def api_create_grid_bot(body: GridBotCreate):
    try:
        bot_id = start_grid_bot(
            symbol=body.symbol,
            investment=body.investment,
            grid_count=body.grid_count,
            price_low=body.price_low,
            price_high=body.price_high,
            mode=body.mode,
            use_base_balance=body.use_base_balance,
        )
        return {"ok": True, "id": bot_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/grid-bots/preview")
def api_grid_preview(symbol: str, investment: float, mode: str = "normal", grid_count: int | None = None):
    """Return auto-calculated (or user-specified) range and grid count without creating a bot.
    Also returns free_usdt so the UI can warn when investment exceeds available balance."""
    try:
        client = MEXCClient()
        sym = symbol.upper()
        if not sym.endswith("USDT"):
            sym += "USDT"
        price = client.get_price(sym)
        low, high = calculate_grid_range(client, sym)
        count = grid_count if (grid_count and grid_count >= 2) else calculate_grid_count(investment, low, high)
        step  = round((high - low) / count, 8)
        profit_per_grid_pct = round(step / low * 100, 4)
        usdt_per_grid = round(investment / count, 2)
        # Fetch free USDT balance so the UI can show an insufficient-funds warning
        try:
            balances = client.get_spot_assets()
            free_usdt = next(
                (float(b.get("free", 0)) for b in balances if b.get("asset", "").upper() == "USDT"),
                0.0,
            )
        except Exception:
            free_usdt = None
        return {
            "symbol":               sym,
            "current_price":        price,
            "price_low":            low,
            "price_high":           high,
            "grid_count":           count,
            "usdt_per_grid":        usdt_per_grid,
            "step":                 step,
            "profit_per_grid_pct":  profit_per_grid_pct,
            "mode":                 mode,
            "est_profit_per_grid":  round(step / price * usdt_per_grid, 4),
            "free_usdt":            round(free_usdt, 2) if free_usdt is not None else None,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/grid-bots/{bot_id}")
def api_get_grid_bot(bot_id: int):
    status = get_grid_bot_status(bot_id)
    if "error" in status and status["error"] == "not found":
        raise HTTPException(status_code=404, detail="Grid bot not found")
    return status


@app.get("/api/grid-bots/{bot_id}/orders")
def api_get_grid_orders(bot_id: int):
    return get_grid_orders(bot_id)


@app.post("/api/grid-bots/{bot_id}/stop")
def api_stop_grid_bot(bot_id: int):
    try:
        stop_grid_bot(bot_id)
        return {"ok": True, "message": "Grid bot stopped"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/grid-bots/{bot_id}/resume")
def api_resume_grid_bot(bot_id: int):
    try:
        resume_grid_bot(bot_id)
        return {"ok": True, "message": "Grid bot resumed"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/grid-bots/{bot_id}")
def api_delete_grid_bot(bot_id: int):
    try:
        stop_grid_bot(bot_id)
        delete_grid_bot(bot_id)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Static file serving – must be LAST so API routes take priority
# ---------------------------------------------------------------------------

@app.get("/{full_path:path}", include_in_schema=False)
def serve_static(full_path: str):
    candidate = os.path.join(_static_dir, full_path)
    if os.path.isfile(candidate):
        if "/_next/static/" in candidate:
            # Hashed assets: immutable, cache forever
            return FileResponse(candidate, headers={"Cache-Control": _IMMUTABLE})
        elif candidate.endswith(".html"):
            return _html_response(candidate)
        else:
            return FileResponse(candidate, headers={"Cache-Control": "public, max-age=3600"})

    # Next.js static export: try path/index.html
    candidate_index = os.path.join(_static_dir, full_path, "index.html")
    if os.path.isfile(candidate_index):
        return _html_response(candidate_index)

    # SPA fallback
    index = os.path.join(_static_dir, "index.html")
    if os.path.exists(index):
        return _html_response(index)
    raise HTTPException(status_code=404)
